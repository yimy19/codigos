from pandas import ExcelWriter
import pandas
from openpyxl import load_workbook

#BETPLAY
data = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='BETPLAY') 
dfdata =  pandas.DataFrame(data)
nuevo1 = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\BETPLAY.xlsx")
dfdata1 =  pandas.DataFrame(nuevo1)    
historico = pandas.concat([dfdata1,dfdata]) 
historico.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\BETPLAY.xlsx" ,sheet_name='BETPLAY')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\BETPLAY.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\BETPLAY.xlsx")
wb.close()


#CARTERA INDIRECTA
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='CARTERA INDIRECTA') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA INDIRECTA.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA INDIRECTA.xlsx",sheet_name='CARTERA INDIRECTA')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA INDIRECTA.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA INDIRECTA.xlsx" )
wb.close()


#CARTERA DIRECTA
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='CARTERA DIRECTA') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA DIRECTA.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA DIRECTA.xlsx",sheet_name='CARTERA DIRECTA')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA DIRECTA.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\CARTERA DIRECTA.xlsx")
wb.close()


#ADULTO MAYOR
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='ADULTO MAYOR') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ADULTO MAYOR.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ADULTO MAYOR.xlsx",sheet_name='ADULTO MAYOR')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ADULTO MAYOR.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ADULTO MAYOR.xlsx")
wb.close()

#LINEA BUSES
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='LINEA BUSES') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\LINEA BUSES.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\LINEA BUSES.xlsx",sheet_name='LINEA BUSES')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\LINEA BUSES.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\LINEA BUSES.xlsx")
wb.close()


#RECARGAS GANA
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='RECARGAS GANA') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RECARGAS GANA.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RECARGAS GANA.xlsx",sheet_name='RECARGAS GANA')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RECARGAS GANA.xlsx")
hoja = wb.active



 
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RECARGAS GANA.xlsx")
wb.close()


#INGRESO SOLIDARIO
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='INGRESO SOLIDARIO') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\INGRESO SOLIDARIO.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\INGRESO SOLIDARIO.xlsx",sheet_name='INGRESO SOLIDARIO')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\INGRESO SOLIDARIO.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\INGRESO SOLIDARIO.xlsx")
wb.close()


#DEVOLUCION DE IVA
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='DEVOLUCION DE IVA') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\DEVOLUCION DE IVA.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\DEVOLUCION DE IVA.xlsx",sheet_name='DEVOLUCION DE IVA')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\DEVOLUCION DE IVA.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\DEVOLUCION DE IVA.xlsx")
wb.close()

#GIROS
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='GIROS') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\GIROS.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\GIROS.xlsx",sheet_name='GIROS')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\GIROS.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\GIROS.xlsx")
wb.close()

#ANULACIÓN GIROS
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='ANULACIÓN GIROS') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN GIROS.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN GIROS.xlsx",sheet_name='ANULACIÓN GIROS')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN GIROS.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN GIROS.xlsx")
wb.close()

#ANULACIÓN EPM
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='ANULACIÓN EPM') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN EPM.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN EPM.xlsx",sheet_name='ANULACIÓN EPM')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN EPM.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\ANULACIÓN EPM.xlsx")
wb.close()

#RIS
data1 = pandas.read_excel(r"D:\Users\gruporeditos\OneDrive - GANA S.A\Auto_Alertas\archivos para el uso\Consolidado automatización de alertasDiarias.xlsx",sheet_name='RIS') 
dfdata3 =  pandas.DataFrame(data1)
nuevo = pandas.read_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RIS.xlsx")
dfdata2 =  pandas.DataFrame(nuevo) 
historico1 = pandas.concat([dfdata2,dfdata3]) 
historico1.to_excel(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RIS.xlsx",sheet_name='RIS')
wb = load_workbook(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RIS.xlsx")
hoja = wb.active
hoja.delete_cols(1)
wb.save(r"D:\\Users\gruporeditos\\OneDrive - GANA S.A\\Auto_Alertas\\historico de productos\\RIS.xlsx")
wb.close()